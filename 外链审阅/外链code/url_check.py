# -*- coding:utf-8 -*-
# 3. 分析综合风险分值大于50分，实际浏览无恶意域名，标记出来，摘取误报分值高的词。
# 4. 分析敏感内容分值大于50分，综合风险分值小于50分的，实际浏览无恶意域名，标记出来，统计是否都恶意。恶意的有多少，非恶意的有多少。
# 5. 给出一定的优化建议
import tkinter as tk
import tkinter.messagebox
import openpyxl
import threading
import requests
from tkinter import filedialog
from selenium import webdriver
# from selenium.webdriver.support.ui import WebDriverWait


class Atuocheck:
    def run(self):
        self.exceldata_get()
        self.chrome_check()

    def threading_it(self, func, *args):
        t = threading.Thread(target=func, args=args)
        t.setDaemon(True)
        t.start()
        # t.join()

    def getStatusCode(self, url):
        r = requests.get(url, allow_redirects=False)
        return r.status_code

    def choose_excel(self):
        # 选择excel文件
        root = tk.Tk()
        root.withdraw()
        filepath = filedialog.askopenfilename()
        return filepath

    def exceldata_get(self):
        global arr_url  # 域名列
        global arr_yesno  # 是否恶意列
        global arr_ps  # 备注列
        global excelpath
        arr_url = []
        arr_yesno = []
        arr_ps = []
        excelpath = self.choose_excel()
        # print(excelpath)
        wb = openpyxl.load_workbook(excelpath)
        sheetnames = wb.get_sheet_names()
        ws = wb.get_sheet_by_name(sheetnames[0])
        '''将域名列表写入文件'''
        # file_handle = open('1.txt', mode='a+')
        # for cell in ws['A']:
        #     file_handle.write(str(cell.value)+"\n")
        # file_handle.close()
        '''将域名列表保存数组'''
        for cell in ws['A']:
            arr_url.append(str(cell.value))
        for cell in ws['B']:
            arr_yesno.append(str(cell.value))
        for cell in ws['C']:
            arr_ps.append(str(cell.value))

    def chrome_check(self):
        del (arr_url[0])
        url_demo = "http://"
        driver = webdriver.Chrome()
        for i in range(0, len(arr_url)):
            print('*****begin*****')
            asw = 0
            full_url = url_demo + str(arr_url[i])
            driver.switch_to.window(driver.window_handles[-1])
            try:
                state = self.getStatusCode(full_url)
                print(state)
                if state in [200, 201, 300, 301]:
                    print(full_url)
                    driver.get(full_url)
                    self.threading_it(self.askbox(i, asw))
                else:
                    new_full_url = url_demo + 'www.' + str(arr_url[i])
                    print(new_full_url)
                    try:
                        driver.get(new_full_url)
                        self.threading_it(self.askbox(i, asw))
                    except Exception:
                        self.threading_it(self.askbox(i, 1))
            except Exception:
                new_full_url = url_demo + 'www.' + str(arr_url[i])
                print(new_full_url)
                try:
                    driver.get(new_full_url)
                    self.threading_it(self.askbox(i, asw))
                except Exception:
                    self.threading_it(self.askbox(i, 1))

            # if state in [200, 201, 300, 301]:
            #     driver.get(full_url)
            #     # abc = WebDriverWait(
            #     #     driver,
            #     #     5).until(lambda e: e.find_element_by_xpath('/html'))
            #     # print('abc:', abc)
            #     # if abc:
            #     #     pass
            #     # else:
            #     #     driver.execute_script('window.stop()')
            #     print(full_url)
            #     self.threading_it(self.askbox(i))
            # else:
            #     new_full_url = url_demo + 'www.' + str(arr_url[i])
            #     driver.get(new_full_url)
            #     print(new_full_url)
            #     self.threading_it(self.askbox(i))

    def askbox(self, i, asw):
        self.i = i
        self.asw = asw
        wb = openpyxl.load_workbook(excelpath)
        sheetnames = wb.get_sheet_names()
        ws = wb.get_sheet_by_name(sheetnames[0])
        if asw == 0:
            win = tk.Tk()
            win.minsize(300, 300)
            result = str(tk.messagebox.askyesno(title='判断',
                                                message='是否为恶意域名？'))
            print(result)
            if result == 'True':
                arr_yesno[i] = '是'
                # 从第二行开始改
                ws.cell(i + 2, 2).value = arr_yesno[i]
            else:
                arr_yesno[i] = '否'
                ws.cell(i + 2, 2).value = arr_yesno[i]
            win.destroy()
        else:
            print('访问失败')
            ws.cell(i + 2, 2).value = '访问失败'
        wb.save(excelpath)
        print('*****end******')

    def exceldata_remake(self):
        pass


if __name__ == "__main__":
    Atuocheck().run()
'''


 '''
