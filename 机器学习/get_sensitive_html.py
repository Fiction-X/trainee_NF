# -*- coding:utf-8 -*-
import re
import time
import openpyxl
import requests
import tkinter as tk
from selenium import webdriver
from tkinter import filedialog
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
# from multiprocessing import Pool


def choose_excel():
    # 选择excel文件
    root = tk.Tk()
    root.withdraw()
    filepath = filedialog.askopenfilename()
    return filepath


def open_excel():
    arr_url_init = []
    excelpath = '敏感表.xlsx'
    # excelpath = choose_excel()
    wb = openpyxl.load_workbook(excelpath, data_only=True)
    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])
    for cell in ws['C']:
        arr_url_init.append(str(cell.value))
    arr_url = list(set(arr_url_init))
    return arr_url


# 获得excel中,没有访问过的url
def get_new_url():
    arr_url = open_excel()
    del_url = []
    with open('history_url.txt', 'r', encoding='utf-8') as f:
        content = f.read()
    content = re.split('\n', content)
    history_url = list(content)
    for i in history_url:
        for n in arr_url:
            if n == i:
                del_url.append(n)
    for i in del_url:
        arr_url.remove(i)
    with open('history_url.txt', 'a', encoding='utf-8') as f:
        for i in arr_url:
            f.write('\n' + i)
    return arr_url


def getstatuscode(url):
    r = requests.get(url, allow_redirects=False)
    return r.status_code


def browser(n, url):
    desired_capabilities = DesiredCapabilities.CHROME
    desired_capabilities["pageLoadStrategy"] = "none"
    chrome_options = Options()
    chrome_options.add_argument('--headless')
    if 'http' in url:
        full_url = url
    else:
        url_demo = "http://"
        full_url = url_demo + url
    print('第', n, '个网站,', 'url:', full_url, '\n')
    driver = webdriver.Chrome(chrome_options=chrome_options)
    # driver = webdriver.Chrome()
    try:
        driver.get(full_url)
        time.sleep(10)
        html = driver.page_source
        nowtime = time.strftime('%m%d%H%M%S', time.localtime(time.time()))
        filetime = str(nowtime)
        print("save in " + str(n) + nowtime)
        f = open("data_html/abnormal/{0}.html".format(str(n) + filetime),
                 'w',
                 encoding='utf-8')
        f.write(html)
        # if len(driver.window_handles) == 1:
        #     pass
        # else:
        #     driver.close()
        # driver.quit()
    except (Exception):
        error = '访问失败'
        print(error)
        f = open("data_html/{0}.html".format(str(url)), 'w', encoding='utf-8')
        f.write(error)
        driver.quit()


# def main():
#     n = 1
#     arr_url = open_excel()
#     # browser(n, 'www.hhoyx.com/')
#     pool = Pool(processes=10)
#     for url in arr_url:
#         pool.apply_async(browser(n, url))
#     pool.close()
#     pool.join()

if __name__ == "__main__":
    url_list = get_new_url()
    print('新增{0}个网址'.format(str(len(url_list))))
    n = 1
    for url in url_list:
        browser(n, url)
        n = n + 1
